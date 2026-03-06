"""Streamlit GUI for Bell Canada BCECN yield spread automater."""

import html
import os
import tempfile

import pandas as pd
import streamlit as st

from excel_writer import append_row, update_charts
from main import BANK_PARSERS, MASTER_FILE, detect_bank, load_env_file

# -- Setup --------------------------------------------------------------------
load_env_file()

st.set_page_config(
    page_title="BCECN Pricing Tool",
    page_icon="\U0001F4C8",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- Theme --------------------------------------------------------------------
st.markdown(
    """
<style>
    #MainMenu, footer {
        visibility: hidden;
    }

    header {
        background: transparent;
    }

    [data-testid="collapsedControl"] {
        margin-top: 0.4rem;
        margin-left: 0.4rem;
        border-radius: 10px;
        background: rgba(17, 27, 46, 0.86);
        border: 1px solid rgba(92, 121, 181, 0.62);
    }

    html, body, [class*="css"] {
        font-family: "Avenir Next", "Trebuchet MS", sans-serif;
    }

    .stApp {
        background:
            radial-gradient(1100px 520px at 8% -10%, rgba(35, 87, 157, 0.24), transparent 52%),
            radial-gradient(900px 480px at 92% 2%, rgba(42, 72, 145, 0.20), transparent 46%),
            linear-gradient(160deg, #050910 0%, #060a14 48%, #04070d 100%);
        color: #e9eef8;
    }

    [data-testid="stAppViewContainer"] > .main .block-container {
        max-width: 1160px;
        padding-top: 1.4rem;
        padding-bottom: 2.4rem;
    }

    .hero-card {
        background: linear-gradient(145deg, rgba(18, 33, 68, 0.88) 0%, rgba(24, 45, 95, 0.76) 100%);
        border: 1px solid rgba(130, 169, 255, 0.26);
        border-radius: 16px;
        padding: 1.5rem 1.7rem;
        box-shadow: 0 20px 50px rgba(5, 12, 30, 0.45);
        margin-bottom: 1.1rem;
    }

    .hero-title {
        color: #f3f7ff;
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
        letter-spacing: 0.2px;
    }

    .hero-subtitle {
        color: #c6d4f1;
        font-size: 1rem;
        margin-bottom: 0.95rem;
    }

    .chip-row {
        display: flex;
        flex-wrap: wrap;
        gap: 0.5rem;
    }

    .chip {
        display: inline-flex;
        align-items: center;
        padding: 0.33rem 0.68rem;
        border-radius: 999px;
        border: 1px solid rgba(123, 161, 247, 0.34);
        background: rgba(23, 40, 78, 0.82);
        color: #e6eeff;
        font-size: 0.82rem;
        line-height: 1.2;
        white-space: nowrap;
    }

    .section-title {
        color: #eff5ff;
        font-size: 1.08rem;
        font-weight: 650;
        margin: 0.2rem 0 0.12rem 0;
    }

    .section-help {
        color: #9db0d8;
        margin-bottom: 0.85rem;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d1629 0%, #101826 100%);
        border-right: 1px solid rgba(116, 140, 194, 0.24);
    }

    [data-testid="stSidebar"] * {
        color: #e4ecff;
    }

    [data-testid="stSidebar"] [data-baseweb="input"] input {
        background: rgba(6, 11, 23, 0.92);
        color: #f0f4ff;
        border: 1px solid rgba(110, 139, 201, 0.42);
        border-radius: 10px;
    }

    .sidebar-heading {
        color: #f4f8ff;
        font-size: 1.12rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }

    .status-pill {
        display: inline-flex;
        align-items: center;
        gap: 0.42rem;
        border-radius: 999px;
        padding: 0.32rem 0.64rem;
        font-size: 0.84rem;
        font-weight: 600;
        margin-top: 0.65rem;
    }

    .status-pill .dot {
        width: 8px;
        height: 8px;
        border-radius: 50%;
    }

    .status-pill.ok {
        background: rgba(52, 211, 153, 0.18);
        border: 1px solid rgba(74, 222, 128, 0.36);
        color: #aff4cc;
    }

    .status-pill.ok .dot {
        background: #4ade80;
    }

    .status-pill.err {
        background: rgba(248, 113, 113, 0.20);
        border: 1px solid rgba(248, 113, 113, 0.48);
        color: #fecaca;
    }

    .status-pill.err .dot {
        background: #f87171;
    }

    [data-testid="stTabs"] button[data-baseweb="tab"] {
        color: #9cb0d8;
        border-radius: 8px 8px 0 0;
        border: 1px solid transparent;
        padding: 0.55rem 0.88rem;
        font-weight: 550;
    }

    [data-testid="stTabs"] button[aria-selected="true"] {
        color: #f3f7ff;
        background: rgba(39, 60, 112, 0.62);
        border-color: rgba(100, 136, 217, 0.62);
    }

    div[data-testid="stFileUploaderDropzone"] {
        background: linear-gradient(180deg, rgba(19, 29, 50, 0.88), rgba(12, 19, 34, 0.9));
        border: 1.5px dashed rgba(103, 136, 207, 0.58);
        border-radius: 12px;
        padding: 1rem;
    }

    div[data-testid="stFileUploaderDropzone"]:hover {
        border-color: rgba(142, 176, 246, 0.84);
    }

    .stButton > button {
        height: 2.72rem;
        border-radius: 10px;
        border: 1px solid rgba(87, 126, 212, 0.96);
        background: linear-gradient(180deg, #2e4f96 0%, #213a6f 100%);
        color: #f6f8ff;
        font-weight: 650;
    }

    .stButton > button:hover {
        border-color: #8fb4ff;
        transform: translateY(-1px);
    }

    .stButton > button:disabled {
        background: #101a30;
        color: #7d8ca8;
        border-color: #2a3450;
        transform: none;
    }

    div[data-testid="stDataFrame"] {
        border-radius: 12px;
        border: 1px solid rgba(83, 108, 160, 0.42);
        overflow: hidden;
    }

    .result-card {
        border-radius: 10px;
        padding: 0.7rem 0.9rem;
        margin: 0.42rem 0 0.7rem 0;
        font-size: 0.94rem;
    }

    .result-ok {
        background: rgba(34, 197, 94, 0.14);
        border: 1px solid rgba(74, 222, 128, 0.36);
        color: #bbf7d0;
    }

    .result-err {
        background: rgba(239, 68, 68, 0.14);
        border: 1px solid rgba(248, 113, 113, 0.45);
        color: #fecaca;
    }

    [data-testid="stProgressBar"] > div > div {
        background: linear-gradient(90deg, #5ca0ff, #7f84ff);
    }

    @media (max-width: 920px) {
        .hero-title {
            font-size: 1.6rem;
        }

        .chip {
            white-space: normal;
        }
    }
</style>
""",
    unsafe_allow_html=True,
)

# -- Helpers ------------------------------------------------------------------
TENORS = ["3Y", "5Y", "7Y", "10Y", "30Y"]
HYBRIDS = ["NC5", "NC10"]
REQUIRED_SHEETS = ("Pricing", "Summary Charts")


def _workbook_status(path: str) -> tuple[str, str, str]:
    """Return workbook readiness status and details for UI display."""
    if not os.path.isfile(path):
        return "err", "Workbook not found", "Set a valid workbook path to enable live writes."

    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
    except Exception as exc:
        return "err", "Cannot open workbook", str(exc)

    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheets]
    if missing:
        return "err", "Missing required sheets", f"Missing: {', '.join(missing)}"

    return "ok", "Workbook ready", "Pricing and Summary Charts sheets detected."


def _build_preview_df(data: dict) -> pd.DataFrame:
    """Build a clean DataFrame from parsed data for display."""
    rows = []
    for ccy in ("CAD", "USD"):
        c = ccy.lower()
        for tenor in TENORS:
            t = tenor.lower()
            spread = data.get(f"{c}_spread_{t}")
            yld = data.get(f"{c}_yield_{t}")
            rows.append(
                {
                    "Currency": ccy,
                    "Tenor": tenor,
                    "Spread (bps)": f"{spread:.1f}" if spread is not None else "-",
                    "Yield (%)": f"{yld:.3%}" if yld is not None else "-",
                }
            )
    for ccy in ("CAD", "USD"):
        c = ccy.lower()
        for hybrid in HYBRIDS:
            h = hybrid.lower()
            spread = data.get(f"{c}_{h}_spread")
            coupon = data.get(f"{c}_{h}_coupon")
            rows.append(
                {
                    "Currency": ccy,
                    "Tenor": hybrid,
                    "Spread (bps)": f"{spread:.1f}" if spread is not None else "-",
                    "Yield (%)": f"{coupon:.3%}" if coupon is not None else "-",
                }
            )
    return pd.DataFrame(rows)


def _process_pdf_file(pdf_path: str, label: str) -> None:
    """Parse one PDF and either preview or write to Excel."""
    try:
        bank_key = detect_bank(pdf_path)
        parser = BANK_PARSERS[bank_key]
        data = parser(pdf_path)
        date_str = data["date"].strftime("%Y-%m-%d")

        if dry_run:
            st.markdown(
                f'<div class="result-card result-ok"><strong>{data["bank"]}</strong> - {date_str} - Preview only</div>',
                unsafe_allow_html=True,
            )
            st.dataframe(_build_preview_df(data), use_container_width=True, hide_index=True)
        else:
            append_row(master_file, data)
            update_charts(master_file)
            st.markdown(
                f'<div class="result-card result-ok"><strong>{data["bank"]}</strong> - {date_str} - Written to workbook</div>',
                unsafe_allow_html=True,
            )
    except Exception as exc:
        safe_label = html.escape(label)
        safe_exc = html.escape(str(exc))
        st.markdown(
            f'<div class="result-card result-err"><strong>{safe_label}</strong> - {safe_exc}</div>',
            unsafe_allow_html=True,
        )


# -- Sidebar ------------------------------------------------------------------
with st.sidebar:
    st.markdown('<div class="sidebar-heading">Control Panel</div>', unsafe_allow_html=True)
    st.caption("Set workbook path and execution mode.")

    master_file = st.text_input(
        "Master workbook path",
        value=os.environ.get("MASTER_FILE", MASTER_FILE),
    )
    dry_run = st.checkbox("Dry run (preview only - no Excel write)", value=False)

    status_class, status_text, status_detail = _workbook_status(master_file)
    st.markdown(
        f'<div class="status-pill {status_class}"><span class="dot"></span>{html.escape(status_text)}</div>',
        unsafe_allow_html=True,
    )
    st.caption(status_detail)

# -- Header -------------------------------------------------------------------
mode_text = "Dry-run preview mode" if dry_run else "Live workbook mode"
status_chip = "Workbook healthy" if status_class == "ok" else "Workbook needs attention"
safe_master = html.escape(master_file)
st.markdown(
    f"""
<div class="hero-card">
    <div class="hero-title">BCECN Pricing Tool</div>
    <div class="hero-subtitle">Bell Canada new issue spread and yield automation</div>
    <div class="chip-row">
        <span class="chip">{mode_text}</span>
        <span class="chip">{status_chip}</span>
        <span class="chip">{safe_master}</span>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

live_mode_blocked = not dry_run and status_class != "ok"
if live_mode_blocked:
    st.markdown(
        '<div class="result-card result-err">Live mode is blocked until workbook validation passes. '
        "Switch to dry run or fix the workbook path/sheets in the sidebar.</div>",
        unsafe_allow_html=True,
    )

# -- Tabs ---------------------------------------------------------------------
tab_pdf, tab_outlook = st.tabs(["\U0001F4C4 Upload PDFs", "\U0001F4E8 Fetch from Outlook"])

# -- Tab 1: Local PDFs --------------------------------------------------------
with tab_pdf:
    st.markdown('<div class="section-title">Local PDF intake</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-help">Upload one or many BCECN PDFs and process them in a single run.</div>',
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader(
        "BCECN PDF files",
        type=["pdf"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        files_df = pd.DataFrame(
            {
                "File": [file.name for file in uploaded_files],
                "Size (MB)": [f"{file.size / (1024 * 1024):.2f}" for file in uploaded_files],
            }
        )
        st.dataframe(files_df, use_container_width=True, hide_index=True)

    col_info, col_btn = st.columns([3.2, 1.2])
    with col_info:
        if uploaded_files:
            st.caption(f"{len(uploaded_files)} file(s) ready.")
        else:
            st.caption("No files selected yet.")
    with col_btn:
        process_clicked = st.button(
            "Process files",
            key="btn_pdf",
            type="primary",
            disabled=not uploaded_files or live_mode_blocked,
            use_container_width=True,
        )

    if process_clicked and uploaded_files:
        progress = st.progress(0, text="Processing files...")
        with tempfile.TemporaryDirectory() as tmp_dir:
            for idx, uploaded_file in enumerate(uploaded_files):
                tmp_path = os.path.join(tmp_dir, uploaded_file.name)
                with open(tmp_path, "wb") as handle:
                    handle.write(uploaded_file.getbuffer())
                _process_pdf_file(tmp_path, uploaded_file.name)
                progress.progress(
                    (idx + 1) / len(uploaded_files),
                    text=f"Processed {idx + 1}/{len(uploaded_files)} file(s)",
                )
        progress.empty()

# -- Tab 2: Outlook fetch -----------------------------------------------------
with tab_outlook:
    st.markdown('<div class="section-title">Outlook fetch</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-help">Pull BCECN attachments from Outlook then run the same parsing pipeline.</div>',
        unsafe_allow_html=True,
    )

    with st.form("outlook_form"):
        left_col, right_col = st.columns(2)
        with left_col:
            email = st.text_input("Email", value=os.environ.get("OUTLOOK_EMAIL", ""))
            server = st.text_input("Server", value=os.environ.get("OUTLOOK_SERVER", "outlook.office365.com"))
        with right_col:
            days = st.number_input(
                "Days back",
                min_value=1,
                value=int(os.environ.get("OUTLOOK_DAYS", "7")),
                step=1,
            )

        sender = st.text_input("Sender filter (optional)", value=os.environ.get("BCECN_SENDER", ""))
        fetch_clicked = st.form_submit_button(
            "Fetch and process",
            type="primary",
            disabled=live_mode_blocked,
            use_container_width=True,
        )

    if fetch_clicked:
        if not email:
            st.markdown(
                '<div class="result-card result-err">Email is required.</div>',
                unsafe_allow_html=True,
            )
        else:
            try:
                from email_fetcher import connect_outlook, fetch_bcecn_pdfs

                auth_placeholder = st.empty()

                def _auth_status(msg):
                    auth_placeholder.info(msg)

                with st.spinner("Authenticating via Microsoft..."):
                    account = connect_outlook(email, server=server, status_callback=_auth_status)

                auth_placeholder.empty()

                with st.spinner("Searching for BCECN emails..."):
                    pdfs = fetch_bcecn_pdfs(
                        account,
                        sender_filter=sender or None,
                        days_back=days,
                    )

                if not pdfs:
                    st.warning("No BCECN PDFs found in the selected range.")
                else:
                    st.caption(f"Found {len(pdfs)} PDF(s). Processing now...")
                    for pdf_path in pdfs:
                        _process_pdf_file(pdf_path, os.path.basename(pdf_path))
            except Exception as exc:
                safe_exc = html.escape(str(exc))
                st.markdown(
                    f'<div class="result-card result-err"><strong>Outlook error</strong> - {safe_exc}</div>',
                    unsafe_allow_html=True,
                )
