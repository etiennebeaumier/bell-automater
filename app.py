#!/usr/bin/env python3
"""BCECN Pricing Tool desktop entry point.

Startup behavior:
- If workbook + preferred PDF source folder are valid and contain parseable PDFs,
  process them immediately and exit.
- Otherwise, launch the full GUI.
"""

import sys
import os

# Ensure the project root is on sys.path (needed when running from PyInstaller bundle)
if getattr(sys, "frozen", False):
    os.chdir(os.path.dirname(sys.executable))
    sys.path.insert(0, os.path.dirname(sys.executable))
else:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import AppConfig, get_default_pdf_dir
from ui.app_window import AppWindow

REQUIRED_SHEETS = {"Pricing"}


def _launch_gui(config: AppConfig) -> None:
    app = AppWindow(config)
    app.mainloop()


def _is_workbook_ready(path: str) -> bool:
    if not path or not os.path.isfile(path):
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
        return REQUIRED_SHEETS.issubset(sheets)
    except Exception:
        return False


def _collect_parseable_pdfs(pdf_dir: str) -> list[str]:
    if not pdf_dir or not os.path.isdir(pdf_dir):
        return []

    pdf_files = sorted(
        os.path.join(pdf_dir, name)
        for name in os.listdir(pdf_dir)
        if name.lower().endswith(".pdf") and os.path.isfile(os.path.join(pdf_dir, name))
    )
    if not pdf_files:
        return []

    from main import detect_bank

    parseable: list[str] = []
    for pdf_path in pdf_files:
        try:
            detect_bank(pdf_path)
            parseable.append(pdf_path)
        except Exception:
            continue
    return parseable


def main():
    """Run startup flow: process PDFs immediately when configured, otherwise open GUI."""
    config = AppConfig()

    master_file = (config.get("master_file") or "").strip()
    preferred_pdf_dir = (config.get("pdf_source_dir") or "").strip() or get_default_pdf_dir()

    if _is_workbook_ready(master_file):
        parseable_pdfs = _collect_parseable_pdfs(preferred_pdf_dir)
        if parseable_pdfs:
            from main import process_many_pdfs

            process_many_pdfs(parseable_pdfs, master_file, dry_run=False)
            return

    _launch_gui(config)


if __name__ == "__main__":
    main()
