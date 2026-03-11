#!/usr/bin/env python3
"""Bell Canada (BCECN) yield spread automater.

Usage:
    # Process a local PDF file:
    python main.py --pdf "BCECN 03.02.26.pdf"

    # Process all PDFs in a directory:
    python main.py --dir ./pdfs/
"""

import argparse
import importlib.util
import os
import sys
from datetime import datetime
from parsers.td import parse_td_pdf
from parsers.scotiabank import parse_scotiabank_pdf
from parsers.cibc import parse_cibc_pdf
from parsers.nbcm import parse_nbcm_pdf
from parsers.bmo import parse_bmo_pdf

MASTER_FILE = os.path.join("data", "Master File.xlsx")
REQUIRED_SHEETS = ("Pricing", "Summary Charts")

BANK_PARSERS = {
    "td": parse_td_pdf,
    "scotiabank": parse_scotiabank_pdf,
    "cibc": parse_cibc_pdf,
    "nbcm": parse_nbcm_pdf,
    "bmo": parse_bmo_pdf,
}


def load_env_file(path: str = ".env") -> None:
    """Load KEY=VALUE pairs from a local .env file into os.environ."""
    if not os.path.exists(path):
        return

    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key:
                continue

            # Strip wrapping quotes for values like PASSWORD="abc123".
            if (value.startswith('"') and value.endswith('"')) or (
                value.startswith("'") and value.endswith("'")
            ):
                value = value[1:-1]

            os.environ.setdefault(key, value)


def _module_available(module_name: str) -> bool:
    """Return True if a Python module can be imported."""
    return importlib.util.find_spec(module_name) is not None


def run_preflight(
    master_file: str,
    require_workbook: bool = True,
    verbose: bool = True,
) -> bool:
    """Validate environment, dependencies, and workbook configuration."""
    failures = 0
    warnings = 0

    def report(level: str, message: str) -> None:
        nonlocal failures, warnings
        if level == "FAIL":
            failures += 1
        elif level == "WARN":
            warnings += 1
        if verbose or level == "FAIL":
            print(f"[{level}] {message}")

    if verbose:
        print("Running preflight checks...")

    if os.path.exists(".env"):
        report("PASS", "Found .env file.")
    else:
        report("WARN", "No .env file found (CLI flags still work). Copy .env.example to .env for easier setup.")

    if _module_available("pdfplumber"):
        report("PASS", "Dependency check: pdfplumber available.")
    else:
        report("FAIL", "Missing dependency: pdfplumber. Install with `pip install pdfplumber`.")

    openpyxl_available = _module_available("openpyxl")
    if openpyxl_available:
        report("PASS", "Dependency check: openpyxl available.")
    elif require_workbook:
        report("FAIL", "Missing dependency: openpyxl. Install with `pip install openpyxl`.")
    else:
        report("WARN", "openpyxl is missing. Install with `pip install openpyxl` before non-dry-run writes.")

    if require_workbook:
        if not os.path.exists(master_file):
            report("FAIL", f"Master workbook not found: {master_file}")
        elif not openpyxl_available:
            report("FAIL", "Cannot validate workbook sheets because openpyxl is unavailable.")
        else:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(master_file, read_only=True, data_only=True)
                missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
                wb.close()

                if missing_sheets:
                    report(
                        "FAIL",
                        f"Workbook is missing required sheets: {', '.join(missing_sheets)}",
                    )
                else:
                    report("PASS", f"Workbook check passed: found {', '.join(REQUIRED_SHEETS)} sheets.")
            except Exception as exc:
                report("FAIL", f"Failed to open workbook '{master_file}': {exc}")
    else:
        report("PASS", "Workbook validation skipped (dry-run mode).")

    if verbose or failures:
        if failures:
            print(f"Preflight failed with {failures} error(s) and {warnings} warning(s).")
        else:
            print(f"Preflight passed with {warnings} warning(s).")

    return failures == 0


def detect_bank(pdf_path: str) -> str:
    """Detect which bank sent the PDF based on filename, then PDF content."""
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")

    filename = pdf_path.lower()

    # Filename-based detection
    if "scotiabank" in filename or "bns" in filename:
        return "scotiabank"
    if "cibc" in filename:
        return "cibc"
    if "nbcm" in filename or "national bank" in filename:
        return "nbcm"
    if "bmo" in filename:
        return "bmo"
    if "td" in filename:
        return "td"

    # Content-based detection as fallback
    import pdfplumber

    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            raise ValueError(f"PDF appears empty: {pdf_path}")
        text = (pdf.pages[0].extract_text() or "").lower()

    if "scotiabank" in text or "bns-internal" in text:
        return "scotiabank"
    if "cibc capital markets" in text or "cibc" in text:
        return "cibc"
    if "national bank" in text or "nbcm" in text:
        return "nbcm"
    if "bmo nesbitt burns" in text or "bmo capital markets" in text:
        return "bmo"
    if "td securities" in text:
        return "td"

    raise ValueError(f"Could not detect bank for: {pdf_path}")


def _format_preview_value(key: str, value) -> str:
    """Format parsed values for dry-run preview output."""
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        if "yield" in key or "coupon" in key:
            return f"{value:.3%}"
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def print_dry_run_preview(pdf_path: str, data: dict) -> None:
    """Print a user-friendly preview of parsed values without writing Excel."""
    ordered_keys = []
    for currency in ("cad", "usd"):
        for metric in ("spread", "yield"):
            for tenor in ("3y", "5y", "7y", "10y", "30y"):
                ordered_keys.append(f"{currency}_{metric}_{tenor}")
    for currency in ("cad", "usd"):
        for nc in ("nc5", "nc10"):
            ordered_keys.append(f"{currency}_{nc}_spread")
            ordered_keys.append(f"{currency}_{nc}_coupon")

    metric_keys = [k for k in data.keys() if k not in {"date", "bank"}]
    extras = sorted(k for k in metric_keys if k not in ordered_keys)
    preview_keys = [k for k in ordered_keys if k in data] + extras

    date_val = data.get("date")
    if isinstance(date_val, datetime):
        date_text = date_val.strftime("%Y-%m-%d")
    else:
        date_text = str(date_val)

    print("\n" + "=" * 72)
    print(f"DRY RUN PREVIEW: {pdf_path}")
    print(f"Bank: {data.get('bank', 'Unknown')} | Date: {date_text}")
    print("-" * 72)
    for key in preview_keys:
        print(f"{key:<20} {_format_preview_value(key, data.get(key))}")
    print("=" * 72)


def process_pdf(pdf_path: str, master_file: str = MASTER_FILE, dry_run: bool = False):
    """Parse one PDF and append its row to the workbook.

    In non-dry-run mode, this function only performs the row append.
    End-of-run deduplication and chart rebuilding are handled by
    `process_many_pdfs` so they run once per batch.
    """
    bank_key = detect_bank(pdf_path)
    parser = BANK_PARSERS.get(bank_key)
    if not parser:
        raise ValueError(f"No parser available for bank: {bank_key}")

    print(f"Parsing {pdf_path} with {bank_key.upper()} parser...")
    data = parser(pdf_path)

    if dry_run:
        print_dry_run_preview(pdf_path, data)
        return data

    from excel_writer import append_row

    print(f"Writing to {master_file}...")
    append_row(master_file, data)
    print("Write complete.")
    return data


def process_many_pdfs(
    pdf_paths: list[str],
    master_file: str,
    dry_run: bool = False,
    avg_start_year: int | None = None,
    avg_end_year: int | None = None,
) -> bool:
    """Process PDFs and run one post-processing step at the end.

    Non-dry-run behavior:
    1) Parse and append rows for each successful PDF.
    2) Once all files are attempted, deduplicate Pricing rows by
       (date, bank case-insensitive), keeping newest rows.
    3) Rebuild Summary Charts once from the final workbook state,
       optionally bounded by `avg_start_year` / `avg_end_year`.
    """
    success = 0
    failed = 0

    for pdf_path in pdf_paths:
        try:
            process_pdf(pdf_path, master_file, dry_run=dry_run)
            success += 1
        except Exception as exc:
            failed += 1
            print(f"Error processing {pdf_path}: {exc}")

    if not dry_run and success > 0:
        from excel_writer import deduplicate_pricing_rows, update_charts

        try:
            removed = deduplicate_pricing_rows(master_file)
            if avg_start_year is None and avg_end_year is None:
                update_charts(master_file)
            else:
                update_charts(
                    master_file,
                    avg_start_year=avg_start_year,
                    avg_end_year=avg_end_year,
                )
            print(
                "Post-processing: "
                f"removed {removed} duplicate row(s) and rebuilt Summary Charts."
            )
        except Exception as exc:
            failed += 1
            print(f"Post-processing failed: {exc}")

    print(f"\nRun summary: {success} succeeded, {failed} failed.")
    return failed == 0


def prompt_with_default(prompt: str, default: str | None = None, required: bool = False) -> str:
    """Prompt user for input with optional default and required validation."""
    while True:
        suffix = f" [{default}]" if default else ""
        value = input(f"{prompt}{suffix}: ").strip()
        if value:
            return value
        if default:
            return default
        if not required:
            return ""
        print("This value is required.")


def prompt_yes_no(prompt: str, default: bool = False) -> bool:
    """Prompt for yes/no input."""
    hint = "Y/n" if default else "y/N"
    while True:
        response = input(f"{prompt} ({hint}): ").strip().lower()
        if not response:
            return default
        if response in {"y", "yes"}:
            return True
        if response in {"n", "no"}:
            return False
        print("Please answer with y or n.")


def interactive_mode(default_master: str) -> None:
    """Run interactive menu mode for no-argument usage."""
    print("Interactive mode")

    while True:
        print("\nChoose an action:")
        print("1) Process one PDF")
        print("2) Process all PDFs in a directory")
        print("3) Run preflight checks")
        print("4) Exit")

        choice = input("Selection [1-4]: ").strip().lower()
        if choice in {"4", "q", "quit", "exit"}:
            return

        if choice == "1":
            pdf_path = prompt_with_default("PDF path", required=True)
            dry_run = prompt_yes_no("Dry run (preview only)", default=True)
            master_file = prompt_with_default("Master workbook path", default_master, required=not dry_run)
            if not run_preflight(
                master_file=master_file,
                require_workbook=not dry_run,
                verbose=True,
            ):
                if not prompt_yes_no("Preflight failed. Continue anyway", default=False):
                    continue
            process_many_pdfs([pdf_path], master_file, dry_run=dry_run)
            continue

        if choice == "2":
            pdf_dir = prompt_with_default("PDF directory", ".", required=True)
            dry_run = prompt_yes_no("Dry run (preview only)", default=True)
            master_file = prompt_with_default("Master workbook path", default_master, required=not dry_run)
            if not os.path.isdir(pdf_dir):
                print(f"Directory not found: {pdf_dir}")
                continue
            pdf_files = sorted(
                os.path.join(pdf_dir, name)
                for name in os.listdir(pdf_dir)
                if name.lower().endswith(".pdf")
            )
            if not pdf_files:
                print(f"No PDF files found in {pdf_dir}")
                continue
            if not run_preflight(
                master_file=master_file,
                require_workbook=not dry_run,
                verbose=True,
            ):
                if not prompt_yes_no("Preflight failed. Continue anyway", default=False):
                    continue
            process_many_pdfs(pdf_files, master_file, dry_run=dry_run)
            continue

        if choice == "3":
            dry_run = prompt_yes_no("Assume dry-run mode (skip workbook validation)", default=False)
            master_file = prompt_with_default("Master workbook path", default_master, required=not dry_run)

            run_preflight(
                master_file=master_file,
                require_workbook=not dry_run,
                verbose=True,
            )
            continue

        print("Invalid selection. Please choose 1-4.")


def main():
    load_env_file(".env")

    default_master = os.environ.get("MASTER_FILE", MASTER_FILE)

    parser = argparse.ArgumentParser(description="BCECN Yield Spread Automater")
    parser.add_argument("--pdf", help="Path to a single PDF file to process")
    parser.add_argument("--dir", help="Directory containing PDF files to process")
    parser.add_argument("--master", default=default_master, help="Path to Master File.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Parse and preview data without writing to Excel")
    parser.add_argument("--check", action="store_true", help="Run preflight checks before processing")
    parser.add_argument("--interactive", action="store_true", help="Launch interactive mode menu")
    args = parser.parse_args()

    operation_requested = bool(args.pdf or args.dir)

    if args.interactive or (not operation_requested and not args.check):
        interactive_mode(default_master=default_master)
        return

    if args.check:
        preflight_ok = run_preflight(
            master_file=args.master,
            require_workbook=not args.dry_run,
            verbose=True,
        )
        if not preflight_ok:
            sys.exit(1)
        if not operation_requested:
            return

    if operation_requested and not args.check:
        preflight_ok = run_preflight(
            master_file=args.master,
            require_workbook=not args.dry_run,
            verbose=False,
        )
        if not preflight_ok:
            print("Run with --check for full diagnostics.")
            sys.exit(1)

    if args.pdf:
        if not process_many_pdfs([args.pdf], args.master, dry_run=args.dry_run):
            sys.exit(1)

    elif args.dir:
        pdf_dir = args.dir
        if not os.path.isdir(pdf_dir):
            print(f"Directory not found: {pdf_dir}")
            sys.exit(1)
        pdf_files = sorted(
            os.path.join(pdf_dir, f)
            for f in os.listdir(pdf_dir)
            if f.lower().endswith(".pdf")
        )
        if not pdf_files:
            print(f"No PDF files found in {pdf_dir}")
            sys.exit(1)
        if not process_many_pdfs(pdf_files, args.master, dry_run=args.dry_run):
            sys.exit(1)

    else:
        interactive_mode(default_master=default_master)


if __name__ == "__main__":
    main()
