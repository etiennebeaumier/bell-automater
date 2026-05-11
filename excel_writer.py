"""Write parsed PDF data into the Master File Excel workbook."""

from copy import copy
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Column mapping: Master File column letters -> parsed data keys
# Pricing sheet columns (1-indexed):
#  A=1 Date, B=2 Bank,
#  C=3..L=12: CAD 3Y-30Y (spread, yield alternating)
#  M=13..V=22: USD 3Y-30Y (spread, yield alternating)
#  W=23..Z=26: CAD NC5/NC10 (spread, coupon)
#  AA=27..AD=30: USD NC5/NC10 (spread, coupon)

COLUMN_MAP = {
    3: "cad_spread_3y", 4: "cad_yield_3y",
    5: "cad_spread_5y", 6: "cad_yield_5y",
    7: "cad_spread_7y", 8: "cad_yield_7y",
    9: "cad_spread_10y", 10: "cad_yield_10y",
    11: "cad_spread_30y", 12: "cad_yield_30y",
    13: "usd_spread_3y", 14: "usd_yield_3y",
    15: "usd_spread_5y", 16: "usd_yield_5y",
    17: "usd_spread_7y", 18: "usd_yield_7y",
    19: "usd_spread_10y", 20: "usd_yield_10y",
    21: "usd_spread_30y", 22: "usd_yield_30y",
    23: "cad_nc5_spread", 24: "cad_nc5_coupon",
    25: "cad_nc10_spread", 26: "cad_nc10_coupon",
    27: "usd_nc5_spread", 28: "usd_nc5_coupon",
    29: "usd_nc10_spread", 30: "usd_nc10_coupon",
}

def _is_macro_enabled(path: str) -> bool:
    """Return True when the workbook extension indicates VBA content."""
    return path.lower().endswith(".xlsm")


def append_row(master_file_path: str, data: dict):
    """Append a new row to the Pricing sheet of the Master File."""
    wb = load_workbook(master_file_path, keep_vba=_is_macro_enabled(master_file_path))
    ws = wb["Pricing"]

    next_row = ws.max_row + 1

    # Find actual next empty row (max_row can overcount with formatting)
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            next_row = r
            break

    center = Alignment(horizontal="center", vertical="center")

    cell = ws.cell(row=next_row, column=1, value=data["date"])
    cell.number_format = "YYYY-MM-DD"
    cell.alignment = center

    cell = ws.cell(row=next_row, column=2, value=data["bank"])
    cell.alignment = center

    # Columns that hold yield/coupon percentages (even-numbered in senior, coupon in hybrid)
    pct_columns = {4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30}

    for col, key in COLUMN_MAP.items():
        value = data.get(key)
        if value is not None:
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.alignment = center
            if col in pct_columns:
                cell.number_format = "0.000%"

    wb.save(master_file_path)
    print(f"Wrote {data['bank']} data for {data['date'].strftime('%Y-%m-%d')} to row {next_row}")
    return wb


def _normalize_date(value):
    """Coerce workbook cell values into `date` objects when possible."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(raw).date()
        except ValueError:
            return None
    return None


def _normalize_bank(value):
    """Normalize a bank label for downstream grouping and comparisons."""
    if value is None:
        return None
    bank = str(value).strip()
    return bank or None


def _collect_pricing_rows(ws_data):
    """Collect rows that have both a usable date and a bank name."""
    rows = []
    for r in range(2, ws_data.max_row + 1):
        date_val = _normalize_date(ws_data.cell(row=r, column=1).value)
        bank_val = _normalize_bank(ws_data.cell(row=r, column=2).value)
        if not date_val or not bank_val:
            continue
        rows.append({"row": r, "date": date_val, "bank": bank_val})
    return rows


def _row_dedup_key(row):
    """Build the case-insensitive `(date, bank)` key used for deduping."""
    return row["date"], row["bank"].casefold()


def _dedupe_rows_by_date_bank(rows):
    """Keep only the newest row for each (date, bank) key."""
    seen = set()
    deduped_reverse = []
    for row in reversed(rows):
        key = _row_dedup_key(row)
        if key in seen:
            continue
        seen.add(key)
        deduped_reverse.append(row)
    return list(reversed(deduped_reverse))


def _snapshot_cell_payload(cell):
    """Capture value and formatting so rows can be reordered losslessly."""
    return {
        "value": cell.value,
        "style": copy(cell._style),
        "comment": copy(cell.comment),
        "hyperlink": copy(cell.hyperlink),
    }


def _restore_cell_payload(cell, payload):
    """Restore a cell payload captured by `_snapshot_cell_payload`."""
    cell.value = payload["value"]
    cell._style = copy(payload["style"])
    cell.comment = copy(payload["comment"])
    cell.hyperlink = copy(payload["hyperlink"])


def _collect_pricing_row_payloads(ws_data):
    """Read complete row payloads, including style metadata, for rewrites."""
    max_col = ws_data.max_column
    rows = []
    for r in range(2, ws_data.max_row + 1):
        cells = [_snapshot_cell_payload(ws_data.cell(row=r, column=c)) for c in range(1, max_col + 1)]
        date_val = _normalize_date(cells[0]["value"]) if cells else None
        bank_val = _normalize_bank(cells[1]["value"]) if len(cells) > 1 else None
        key = (date_val, bank_val.casefold()) if date_val and bank_val else None
        rows.append(
            {
                "row": r,
                "date": date_val,
                "bank": bank_val,
                "dedup_key": key,
                "cells": cells,
            }
        )
    return rows


def _dedupe_pricing_row_payloads(rows):
    """Remove duplicate payload rows while keeping the newest copy."""
    seen = set()
    deduped_reverse = []
    removed = 0

    for row in reversed(rows):
        key = row["dedup_key"]
        if key is None:
            deduped_reverse.append(row)
            continue
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        deduped_reverse.append(row)

    return list(reversed(deduped_reverse)), removed


def _sort_pricing_row_payloads(rows):
    """Sort valid rows by bank/date and keep incomplete rows at the bottom."""
    valid_rows = [row for row in rows if row["dedup_key"] is not None]
    invalid_rows = [row for row in rows if row["dedup_key"] is None]
    sorted_valid = sorted(valid_rows, key=lambda row: (row["bank"].casefold(), row["date"], row["row"]))
    return sorted_valid + invalid_rows


def _rewrite_pricing_rows(ws_data, rows):
    """Rewrite Pricing rows from a normalized in-memory payload list."""
    existing_rows = ws_data.max_row - 1
    if existing_rows > 0:
        ws_data.delete_rows(2, existing_rows)

    for target_row, row_payload in enumerate(rows, start=2):
        for col, cell_payload in enumerate(row_payload["cells"], start=1):
            cell = ws_data.cell(row=target_row, column=col)
            _restore_cell_payload(cell, cell_payload)


def deduplicate_pricing_rows(master_file_path: str) -> int:
    """Deduplicate and reorder Pricing rows.

    1) Remove duplicate rows by (date, bank case-insensitive), keeping newest.
    2) Order valid rows by bank (case-insensitive), then by date ascending.
    3) Keep rows with missing date or bank at the bottom in original order.

    Returns the number of duplicates removed.
    """
    wb = load_workbook(master_file_path, keep_vba=_is_macro_enabled(master_file_path))
    ws_data = wb["Pricing"]

    all_rows = _collect_pricing_row_payloads(ws_data)
    deduped_rows, removed = _dedupe_pricing_row_payloads(all_rows)
    sorted_rows = _sort_pricing_row_payloads(deduped_rows)
    order_changed = [row["row"] for row in sorted_rows] != [row["row"] for row in deduped_rows]

    if removed or order_changed:
        _rewrite_pricing_rows(ws_data, sorted_rows)
        wb.save(master_file_path)

    return removed


