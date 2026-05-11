from datetime import date, datetime
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import excel_writer


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


class DeduplicatePricingRowsTests(unittest.TestCase):
    def _create_workbook(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pricing"
        ws.cell(row=1, column=1, value="Date")
        ws.cell(row=1, column=2, value="Bank")
        ws.cell(row=1, column=3, value="CAD 3Y Spread")
        for row_number, row_values in enumerate(rows, start=2):
            ws.cell(row=row_number, column=1, value=row_values["date"])
            ws.cell(row=row_number, column=2, value=row_values["bank"])
            ws.cell(row=row_number, column=3, value=row_values["spread"])
        wb.save(path)

    def _read_pricing_rows(self, path):
        wb = load_workbook(path)
        ws = wb["Pricing"]
        rows = []
        for row_number in range(2, ws.max_row + 1):
            rows.append(
                (
                    _to_date(ws.cell(row=row_number, column=1).value),
                    ws.cell(row=row_number, column=2).value,
                    ws.cell(row=row_number, column=3).value,
                )
            )
        return rows

    def test_duplicate_date_bank_keeps_newest_case_insensitive(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = f"{temp_dir}/master.xlsx"
            self._create_workbook(
                workbook_path,
                rows=[
                    {"date": date(2024, 1, 2), "bank": "BankA", "spread": 100},
                    {"date": date(2024, 1, 2), "bank": "banka", "spread": 999},
                    {"date": date(2024, 1, 3), "bank": "BankB", "spread": 120},
                    {"date": date(2024, 1, 3), "bank": "BANKB", "spread": 130},
                ],
            )

            removed = excel_writer.deduplicate_pricing_rows(workbook_path)
            self.assertEqual(removed, 2)

            self.assertEqual(
                self._read_pricing_rows(workbook_path),
                [
                    (date(2024, 1, 2), "banka", 999),
                    (date(2024, 1, 3), "BANKB", 130),
                ],
            )

    def test_non_duplicate_rows_are_sorted_by_bank_then_date(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = f"{temp_dir}/master.xlsx"
            rows = [
                {"date": date(2024, 1, 2), "bank": "BankB", "spread": 90},
                {"date": date(2024, 1, 3), "bank": "BankA", "spread": 110},
                {"date": date(2024, 1, 2), "bank": "BankA", "spread": 100},
            ]
            self._create_workbook(workbook_path, rows=rows)

            removed = excel_writer.deduplicate_pricing_rows(workbook_path)

            self.assertEqual(removed, 0)
            self.assertEqual(
                self._read_pricing_rows(workbook_path),
                [
                    (date(2024, 1, 2), "BankA", 100),
                    (date(2024, 1, 3), "BankA", 110),
                    (date(2024, 1, 2), "BankB", 90),
                ],
            )

    def test_rows_missing_date_or_bank_are_untouched(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = f"{temp_dir}/master.xlsx"
            self._create_workbook(
                workbook_path,
                rows=[
                    {"date": date(2024, 1, 2), "bank": "BankA", "spread": 100},
                    {"date": date(2024, 1, 2), "bank": "banka", "spread": 999},
                    {"date": None, "bank": "NoDate", "spread": 10},
                    {"date": date(2024, 1, 4), "bank": None, "spread": 11},
                ],
            )

            removed = excel_writer.deduplicate_pricing_rows(workbook_path)
            self.assertEqual(removed, 1)
            self.assertEqual(
                self._read_pricing_rows(workbook_path),
                [
                    (date(2024, 1, 2), "banka", 999),
                    (None, "NoDate", 10),
                    (date(2024, 1, 4), None, 11),
                ],
            )

    def test_rows_missing_date_or_bank_move_to_bottom_in_original_order(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = f"{temp_dir}/master.xlsx"
            self._create_workbook(
                workbook_path,
                rows=[
                    {"date": None, "bank": "NoDate", "spread": 10},
                    {"date": date(2024, 1, 2), "bank": "BankB", "spread": 90},
                    {"date": date(2024, 1, 2), "bank": "BankA", "spread": 100},
                    {"date": date(2024, 1, 4), "bank": None, "spread": 11},
                ],
            )

            removed = excel_writer.deduplicate_pricing_rows(workbook_path)
            self.assertEqual(removed, 0)
            self.assertEqual(
                self._read_pricing_rows(workbook_path),
                [
                    (date(2024, 1, 2), "BankA", 100),
                    (date(2024, 1, 2), "BankB", 90),
                    (None, "NoDate", 10),
                    (date(2024, 1, 4), None, 11),
                ],
            )

    def test_reorder_preserves_cell_style_metadata(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = f"{temp_dir}/master.xlsx"
            self._create_workbook(
                workbook_path,
                rows=[
                    {"date": date(2024, 1, 2), "bank": "BankB", "spread": 90},
                    {"date": date(2024, 1, 2), "bank": "BankA", "spread": 100},
                ],
            )

            wb = load_workbook(workbook_path)
            ws = wb["Pricing"]
            ws.cell(row=2, column=3).fill = PatternFill(fill_type="solid", fgColor="00FF0000")
            ws.cell(row=2, column=3).number_format = "0.000"
            ws.cell(row=3, column=3).fill = PatternFill(fill_type="solid", fgColor="0000FF00")
            ws.cell(row=3, column=3).number_format = "0.00"
            wb.save(workbook_path)

            removed = excel_writer.deduplicate_pricing_rows(workbook_path)
            self.assertEqual(removed, 0)

            wb = load_workbook(workbook_path)
            ws = wb["Pricing"]

            self.assertEqual(ws.cell(row=2, column=2).value, "BankA")
            self.assertEqual(ws.cell(row=2, column=3).fill.fgColor.rgb, "0000FF00")
            self.assertEqual(ws.cell(row=2, column=3).number_format, "0.00")

            self.assertEqual(ws.cell(row=3, column=2).value, "BankB")
            self.assertEqual(ws.cell(row=3, column=3).fill.fgColor.rgb, "00FF0000")
            self.assertEqual(ws.cell(row=3, column=3).number_format, "0.000")


if __name__ == "__main__":
    unittest.main()
