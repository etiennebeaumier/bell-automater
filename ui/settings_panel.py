"""Settings sidebar panel."""

import os
from datetime import datetime
from tkinter import filedialog
import customtkinter as ctk
from config import get_default_master_file


class SettingsPanel(ctk.CTkFrame):
    def __init__(self, master, config, on_settings_changed=None, **kwargs):
        super().__init__(master, width=280, **kwargs)
        self.config = config
        self.on_settings_changed = on_settings_changed

        self.grid_columnconfigure(0, weight=1)

        # -- Title --
        title = ctk.CTkLabel(self, text="Control Panel", font=ctk.CTkFont(size=16, weight="bold"))
        title.grid(row=0, column=0, padx=16, pady=(16, 4), sticky="w")

        subtitle = ctk.CTkLabel(self, text="Workbook path and execution mode", font=ctk.CTkFont(size=12), text_color="gray")
        subtitle.grid(row=1, column=0, padx=16, pady=(0, 12), sticky="w")

        # -- Workbook path --
        wb_label = ctk.CTkLabel(self, text="Master Workbook", font=ctk.CTkFont(size=13, weight="bold"))
        wb_label.grid(row=2, column=0, padx=16, pady=(8, 2), sticky="w")

        path_frame = ctk.CTkFrame(self, fg_color="transparent")
        path_frame.grid(row=3, column=0, padx=16, pady=(0, 4), sticky="ew")
        path_frame.grid_columnconfigure(0, weight=1)

        configured_workbook = config.get("master_file", "")
        default_workbook = configured_workbook or get_default_master_file()
        self.workbook_var = ctk.StringVar(value=default_workbook)
        self.workbook_entry = ctk.CTkEntry(path_frame, textvariable=self.workbook_var, placeholder_text="Select workbook...")
        self.workbook_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        browse_btn = ctk.CTkButton(path_frame, text="Browse", width=70, command=self._browse_workbook)
        browse_btn.grid(row=0, column=1)

        # -- Status indicator --
        self.status_label = ctk.CTkLabel(self, text="", font=ctk.CTkFont(size=12))
        self.status_label.grid(row=4, column=0, padx=16, pady=(4, 8), sticky="w")
        self._update_status()

        # -- Year range for averaged spread charts --
        years_label = ctk.CTkLabel(self, text="Average Spread Year Range", font=ctk.CTkFont(size=13, weight="bold"))
        years_label.grid(row=5, column=0, padx=16, pady=(10, 2), sticky="w")

        years_frame = ctk.CTkFrame(self, fg_color="transparent")
        years_frame.grid(row=6, column=0, padx=16, pady=(0, 8), sticky="ew")
        years_frame.grid_columnconfigure(0, weight=1)
        years_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(years_frame, text="Start Year", font=ctk.CTkFont(size=12), text_color="gray").grid(
            row=0, column=0, padx=(0, 4), pady=(0, 2), sticky="w"
        )
        ctk.CTkLabel(years_frame, text="End Year", font=ctk.CTkFont(size=12), text_color="gray").grid(
            row=0, column=1, padx=(4, 0), pady=(0, 2), sticky="w"
        )

        fallback_year = str(datetime.now().year)
        self.start_year_menu = ctk.CTkOptionMenu(
            years_frame,
            values=[fallback_year],
            command=lambda _choice: self._on_change(),
        )
        self.start_year_menu.grid(row=1, column=0, padx=(0, 4), sticky="ew")

        self.end_year_menu = ctk.CTkOptionMenu(
            years_frame,
            values=[fallback_year],
            command=lambda _choice: self._on_change(),
        )
        self.end_year_menu.grid(row=1, column=1, padx=(4, 0), sticky="ew")

        # -- Dry run checkbox --
        self.dry_run_var = ctk.BooleanVar(value=config.get("dry_run", False))
        self.dry_run_check = ctk.CTkCheckBox(
            self, text="Dry run (preview only)", variable=self.dry_run_var,
            command=self._on_change,
        )
        self.dry_run_check.grid(row=7, column=0, padx=16, pady=(8, 8), sticky="w")

        # -- Appearance mode --
        theme_label = ctk.CTkLabel(self, text="Appearance", font=ctk.CTkFont(size=13, weight="bold"))
        theme_label.grid(row=8, column=0, padx=16, pady=(16, 2), sticky="w")

        self.theme_menu = ctk.CTkOptionMenu(
            self, values=["Dark", "Light", "System"],
            command=self._change_theme,
        )
        self.theme_menu.set(config.get("appearance_mode", "dark").capitalize())
        self.theme_menu.grid(row=9, column=0, padx=16, pady=(0, 8), sticky="ew")

        # -- Save settings button --
        save_btn = ctk.CTkButton(self, text="Save Settings", command=self._save_settings)
        save_btn.grid(row=10, column=0, padx=16, pady=(16, 16), sticky="ew")

        self._refresh_year_options(prefer_config=True)

        # Bind entry change
        self.workbook_var.trace_add("write", lambda *_: self._on_workbook_change())

    def _browse_workbook(self):
        seed_path = self.workbook_var.get() or get_default_master_file()
        initial_dir = os.path.dirname(seed_path) if seed_path else os.path.expanduser("~")
        path = filedialog.askopenfilename(
            title="Select Master Workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
            initialdir=initial_dir,
            initialfile=os.path.basename(seed_path) if seed_path else "",
        )
        if path:
            self.workbook_var.set(path)

    def _on_workbook_change(self):
        self._update_status()
        self._refresh_year_options(prefer_config=False)
        self._on_change()

    def _update_status(self):
        path = self.workbook_var.get()
        if not path:
            self.status_label.configure(text="  No workbook selected", text_color="#f87171")
            return
        if not os.path.isfile(path):
            self.status_label.configure(text="  Workbook not found", text_color="#f87171")
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            sheets = set(wb.sheetnames)
            wb.close()
            required = {"Pricing", "Summary Charts"}
            missing = required - sheets
            if missing:
                self.status_label.configure(text=f"  Missing: {', '.join(missing)}", text_color="#f87171")
            else:
                self.status_label.configure(text="  Workbook ready", text_color="#4ade80")
        except Exception as e:
            self.status_label.configure(text=f"  Error: {e}", text_color="#f87171")

    def _load_available_years(self) -> list[int]:
        path = self.workbook_var.get()
        current_year = datetime.now().year
        if not path or not os.path.isfile(path):
            return [current_year]

        years = set()
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            if "Pricing" in wb.sheetnames:
                ws = wb["Pricing"]
                for r in range(2, ws.max_row + 1):
                    date_val = ws.cell(row=r, column=1).value
                    if hasattr(date_val, "year"):
                        years.add(int(date_val.year))
                        continue
                    if isinstance(date_val, str):
                        raw = date_val.strip()
                        if not raw:
                            continue
                        parsed = None
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
                            try:
                                parsed = datetime.strptime(raw, fmt)
                                break
                            except ValueError:
                                continue
                        if parsed is None:
                            try:
                                parsed = datetime.fromisoformat(raw)
                            except ValueError:
                                parsed = None
                        if parsed is not None:
                            years.add(parsed.year)
            wb.close()
        except Exception:
            return [current_year]

        if not years:
            return [current_year]
        return sorted(years)

    @staticmethod
    def _parse_year(raw, fallback: int) -> int:
        try:
            return int(raw)
        except (TypeError, ValueError):
            return fallback

    def _refresh_year_options(self, prefer_config: bool):
        years = self._load_available_years()
        year_values = [str(y) for y in years]

        self.start_year_menu.configure(values=year_values)
        self.end_year_menu.configure(values=year_values)

        if prefer_config:
            start_pref = self._parse_year(self.config.get("avg_start_year"), years[0])
            end_pref = self._parse_year(self.config.get("avg_end_year"), years[-1])
        else:
            start_pref = self.avg_start_year
            end_pref = self.avg_end_year

        start_year = start_pref if start_pref in years else years[0]
        end_year = end_pref if end_pref in years else years[-1]

        self.start_year_menu.set(str(start_year))
        self.end_year_menu.set(str(end_year))

    def _change_theme(self, choice):
        ctk.set_appearance_mode(choice.lower())
        self._on_change()

    def _on_change(self):
        if self.on_settings_changed:
            self.on_settings_changed()

    def _save_settings(self):
        self.config["master_file"] = self.workbook_var.get()
        self.config["dry_run"] = self.dry_run_var.get()
        self.config["appearance_mode"] = self.theme_menu.get().lower()
        self.config["avg_start_year"] = self.avg_start_year
        self.config["avg_end_year"] = self.avg_end_year
        self.config.save()
        self._update_status()
        self._refresh_year_options(prefer_config=False)

    @property
    def master_file(self) -> str:
        return self.workbook_var.get()

    @property
    def dry_run(self) -> bool:
        return self.dry_run_var.get()

    @property
    def avg_start_year(self) -> int:
        return self._parse_year(self.start_year_menu.get(), datetime.now().year)

    @property
    def avg_end_year(self) -> int:
        return self._parse_year(self.end_year_menu.get(), datetime.now().year)

    def is_workbook_ready(self) -> bool:
        path = self.workbook_var.get()
        if not path or not os.path.isfile(path):
            return False
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            sheets = set(wb.sheetnames)
            wb.close()
            return {"Pricing", "Summary Charts"}.issubset(sheets)
        except Exception:
            return False
